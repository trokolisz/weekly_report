from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from .models import Task
from .forms import TaskForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserChangeForm, PasswordChangeForm, AuthenticationForm
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash, authenticate
from django.contrib.auth import logout as auth_logout

class TaskCreateView(LoginRequiredMixin, CreateView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New Task'
        return context
    
    model = Task
    form_class = TaskForm
    template_name = 'tasks/task_form.html'
    success_url = reverse_lazy('task_list')

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


from django.views.generic.list import ListView
from django.utils import timezone
from datetime import timedelta
from django.contrib.auth import login

class TaskListView(LoginRequiredMixin, ListView):
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Task'
        return context
    model = Task
    template_name = 'tasks/task_list.html'
    context_object_name = 'tasks'

    def get_queryset(self):
        user = self.request.user
        week = self.request.GET.get('week')
        if week:
            week = int(week)
        else:
            week = timezone.now().isocalendar()[1]

        tasks = Task.objects.filter(
            user=user,
            created_at__week=week
        )
        return tasks

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['week'] = self.request.GET.get('week', timezone.now().isocalendar()[1])
        return context



from django.contrib.auth.mixins import UserPassesTestMixin
from django.utils import timezone
from django.views.generic.edit import UpdateView

class TaskUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Task
    form_class = TaskForm
    template_name = 'tasks/task_form.html'
    success_url = reverse_lazy('task_list')

    def test_func(self):
        task = self.get_object()
        one_week_ago = timezone.now() - timedelta(weeks=1)
        return task.user == self.request.user and task.created_at >= one_week_ago



class SubordinateTaskListView(LoginRequiredMixin, ListView):
    model = Task
    template_name = 'tasks/subordinate_task_list.html'
    context_object_name = 'tasks'

    def get_queryset(self):
        user = self.request.user
        subordinates = user.get_all_subordinates()
        week = self.request.GET.get('week')
        if week:
            week = int(week)
        else:
            week = timezone.now().isocalendar()[1]

        tasks = Task.objects.filter(
            user__in=subordinates,
            created_at__week=week
        )
        return tasks

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['week'] = self.request.GET.get('week', timezone.now().isocalendar()[1])
        return context



from django.http import HttpResponse

def export_tasks_text(request):
    tasks = Task.objects.filter(user=request.user)
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename=tasks.txt'

    for task in tasks:
        response.write(f"{task.created_at.date()} - {task.description} ({task.time_spent} mins)\n")

    return response



import openpyxl
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect

def export_tasks_excel(request):
    tasks = Task.objects.filter(user=request.user)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Tasks'

    # Define the titles for columns
    columns = ['Date', 'Description', 'Time Spent (mins)']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all tasks
    for task in tasks:
        row_num += 1
        row = [
            task.created_at.date(),
            task.description,
            task.time_spent,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Adjust column widths
    for col_num in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = 25

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=tasks.xlsx'
    workbook.save(response)
    return response


# Profile Page View
@login_required
def profile_view(request):
    return render(request, 'profile.html', {
        'user': request.user
    })

# Edit Profile Page View
@login_required
def profile_edit_view(request):
    if request.method == 'POST':
        form = UserChangeForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('profile')
    else:
        form = UserChangeForm(instance=request.user)
    
    return render(request, 'profile_edit.html', {
        'form': form
    })

# Logout View
def logout_view(request):
    auth_logout(request)
    return render(request, 'logout.html')

# Change Password View
@login_required
def change_password_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important to keep user logged in after password change
            messages.success(request, 'Your password was successfully updated!')
            return redirect('profile')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'password_change.html', {
        'form': form
    })

def registerView(request):
    return render(request, 'register.html')

def loginView(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Welcome, {username}!')
                return redirect('task_list')  # Redirect to a success page.
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = AuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})

